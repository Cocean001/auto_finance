import os

import openpyxl
import pandas as pd
import xlrd
import sqlite3
from datetime import datetime
from dateutil.relativedelta import relativedelta
conn = sqlite3.connect('finance.db')
cur = conn.cursor()
import numpy as np
from alive_progress import alive_bar
# conn.execute('''
#     CREATE TABLE "匹配情况临时表"(
#         trans_id int,
#         "金额浮动值" int,
#         "可信度" int,
#         "日期匹配度" int,
#         "金额匹配度" int,
#         "总体匹配度" int,
#         "匹配情况" text
#     )
# ''')


# 插入数据
# conn.execute("INSERT INTO transaction_time (trans_id,float_rate,validity) VALUES(1,'11','123456')")
# conn.commit() # 提交数据
# conn.close()


# ———————————— 1. 导入主表及附表 ————————————
main_table_path = "/Users/acroon/Downloads/finance/main.xlsx"
main_table = xlrd.open_workbook(main_table_path)
sheet_main_table = main_table.sheet_by_name("all")
main_rows = sheet_main_table.nrows

sub_table_path = "/Users/acroon/Downloads/finance/sub.xlsx"
sub_table = xlrd.open_workbook(sub_table_path)
sheet_sub_table = sub_table.sheet_by_name("all")
sub_rows = sheet_sub_table.nrows

# 创建路径
tempbook_path_sub = "/Users/acroon/Downloads/finance/temp_sub.xlsx"
tempbook_path_main = "/Users/acroon/Downloads/finance/temp_main.xlsx"
tempbook_path_data = "/Users/acroon/Downloads/finance/temp_data.xlsx"
tempbook_path_data2 = "/Users/acroon/Downloads/finance/temp_data2.xlsx"
tempbook_path_data3 = "/Users/acroon/Downloads/finance/temp_data3.xlsx"

# 创建临时工作簿（数据存储）sub
tempbook_sub = openpyxl.Workbook() # 创建workbook对象
tempbook_sub_act = tempbook_sub.active # 获取默认工作簿
tempbook_sub_act.title = "temp"
tempbook_sub_new_sheet = tempbook_sub.create_sheet(0)
temp_col = ['序号','核对账户','交易日期','交易金额','账户余额','对方户名']

for i in range(0,len(temp_col)):
    j = i+1
    tempbook_sub_act.cell(1,j).value = temp_col[i]
tempbook_sub.save(tempbook_path_sub)

# 创建临时工作簿（数据处理）data
tempbook_data = openpyxl.Workbook() # 创建workbook对象
tempbook_data_act = tempbook_data.active # 获取默认工作簿
tempbook_data_act.title = "temp_data"
tempbook_data_col = ['序号','金额浮动值','可信度','日期匹配度','金额匹配度','总体匹配度']
for i in range(0,len(tempbook_data_col)):
    j = i+1
    tempbook_data_act.cell(1,j).value = tempbook_data_col[i]
tempbook_data.save(tempbook_path_data)

# 创建临时工作簿（数据处理）data2
tempbook_data2 = openpyxl.Workbook() # 创建workbook对象
tempbook_data2_act = tempbook_data2.active # 获取默认工作簿
tempbook_data2_act.title = "new_data"
tempbook_data2.save(tempbook_path_data2)

# 创建临时工作簿（数据处理）data3
tempbook_data3 = openpyxl.Workbook() # 创建workbook对象
tempbook_data3_act = tempbook_data3.active # 获取默认工作簿
tempbook_data3_act.title = "temp_data"
tempbook_data3.save(tempbook_path_data3)

# 创建临时工作簿（主表）main
tempbook_main = openpyxl.Workbook()
tempbook_main_act = tempbook_main.active
tempbook_main_act.title = "main"
tempbook_main_col = ['序号','交易日期','交易金额','单独序列号','匹配情况','总体匹配度','总体可信度','匹配明细']
for i in range(0,len(tempbook_main_col)):
    j = i+1
    tempbook_main_act.cell(1,j).value = tempbook_main_col[i]
tempbook_main.save(tempbook_path_main)

# ———————————— 2. 数据清洗 ————————————————

# 附表序列化
sub_clean_number = []
sub_clean_date = []
sub_clean_money = []
sub_clean_balance = []
sub_clean_card = []
sub_clean_to = []
x = 1
for x in range(1, sub_rows):
    # 序号序列化
    cell_sub_number = sheet_sub_table.cell_value(x, 0)
    cell_sub_number = int(cell_sub_number)
    sub_clean_number.append(cell_sub_number)

    # 日期序列化
    cell_sub_date = sheet_sub_table.cell_value(x, 4)
    # cell_sub_date = xlrd.xldate.xldate_as_tuple(cell_sub_date, 1)
    # if type(cell_sub_date) == 'int':
    #     cell_sub_date = str(int(cell_sub_date))
    # elif type(cell_sub_date) == 'str':
    # new = pd.to_datetime(cell_sub_date)
    new = pd.to_datetime(str(int(cell_sub_date)))
    # print("类型为：",type(cell_sub_date))
    sub_clean_date.append(new)

    # 交易金额序列化
    cell_sub_money = sheet_sub_table.cell_value(x, 5)
    # new = round(cell_sub_money, 3)
    new = int(cell_sub_money)
    sub_clean_money.append(new)

    # 余额序列化
    cell_sub_balance = sheet_sub_table.cell_value(x, 6)
    # new = round(cell_sub_balance, 3)
    new = int(cell_sub_balance)
    sub_clean_balance.append(new)

    # 核对账户序列化
    cell_sub_card = sheet_sub_table.cell_value(x, 9)
    new = str(cell_sub_card)
    sub_clean_card.append(new)

    # 对方账户序列化
    cell_sub_to = sheet_sub_table.cell_value(x, 8)
    new = str(cell_sub_to)
    sub_clean_to.append(new)

# 将序列化结果导入临时表
for i in range(0, len(sub_clean_money)):
    j = i+2
    tempbook_sub_act.cell(j, 1).value = sub_clean_number[i]
    tempbook_sub_act.cell(j, 3).value = sub_clean_date[i]
    tempbook_sub_act.cell(j, 4).value = sub_clean_money[i]
    tempbook_sub_act.cell(j, 5).value = sub_clean_balance[i]
    tempbook_sub_act.cell(j, 2).value = sub_clean_card[i]
    tempbook_sub_act.cell(j, 6).value = sub_clean_card[i]
tempbook_sub.save(tempbook_path_sub)

# 按照时间顺序排列（附表）
# range_date = pd.read_excel(tempbook_path_sub)
# range_date.sort_values(by=['交易日期','核对账户'],inplace=True,ascending=False)
# range_date.to_excel(tempbook_path_sub)

# 主表序列化并写入文件
main_clean_number = []
main_clean_date = []
main_clean_money = []
main_clean_ser = []

x = 1
for x in range(1, main_rows):
    # 序号序列化（主表）
    cell_main_number = sheet_main_table.cell_value(x, 0)
    cell_main_number = int(cell_main_number)
    main_clean_number.append(cell_main_number)

    # 收款日期序列化（主表）
    cell_main_date = sheet_main_table.cell_value(x, 4)
    # new = pd.to_datetime(cell_main_date)
    new = pd.to_datetime(str(int(cell_main_date)))
    main_clean_date.append(new)

    # 金额序列化（主表）
    cell_main_money = sheet_main_table.cell_value(x, 7)
    new = round(cell_main_money, 3)
    main_clean_money.append(new)

    # 唯一号码序列化（主表）
    cell_main_ser = sheet_main_table.cell_value(x, 6)
    cell_main_ser = str(cell_main_ser)
    main_clean_ser.append(cell_main_ser)

for i in range(0, len(main_clean_number)):
    # 写入主表临时表
    j = i+2
    tempbook_main_act.cell(j, 1).value = main_clean_number[i]
    tempbook_main_act.cell(j, 2).value = main_clean_date[i]
    tempbook_main_act.cell(j, 3).value = main_clean_money[i]
    tempbook_main_act.cell(j, 4).value = main_clean_ser[i]
tempbook_main.save(tempbook_path_main)

# 按照时间顺序排列（主表）
# range_date = pd.read_excel(tempbook_path_main)
# range_date.sort_values(by=['交易日期','核对账户'],inplace=True,ascending=False)
# range_date.to_excel(tempbook_path_main)

# ———————————— 3. 匹配日期及金额 ————————————
# 数据初始化：主表提取日期及金额，作为查询条件，进入查询。若查不到，则自动往后查3天。


with alive_bar(main_rows, title="正在匹配流水中", force_tty=True) as bar:
    for xx in range(1,main_rows):
        yy = xx + 1
        try:
            # 读取主表
            main_compare = pd.read_excel(tempbook_path_main)
            main_compare_date = main_compare.iloc[xx]['交易日期']
            main_compare_money = main_compare.iloc[xx]['交易金额']
            sub_compare = pd.read_excel(tempbook_path_sub)
            sub_compare_date_find = sub_compare[(sub_compare['交易日期'] == main_compare_date) & (sub_compare['交易金额'] > 0)]
            df = pd.DataFrame(sub_compare_date_find)

            if df.empty: # 找不到日期，就直接查询后3天
                # 创建向后浮动3天数值，并写入data2
                date_float_1 = main_compare_date + relativedelta(days=1)
                date_float_2 = main_compare_date + relativedelta(days=2)
                date_float_3 = main_compare_date + relativedelta(days=3)
                sub_compare_date_find1 = sub_compare[(sub_compare['交易日期'] == date_float_1) & (sub_compare['交易金额'] > 0)]
                sub_compare_date_find2 = sub_compare[(sub_compare['交易日期'] == date_float_2) & (sub_compare['交易金额'] > 0)]
                sub_compare_date_find3 = sub_compare[(sub_compare['交易日期'] == date_float_3) & (sub_compare['交易金额'] > 0)]
                df1 = pd.DataFrame(sub_compare_date_find1)
                df2 = pd.DataFrame(sub_compare_date_find2)
                df3 = pd.DataFrame(sub_compare_date_find3)

                # 近3天没有交易情况
                if df1.empty and df2.empty and df3.empty:
                    # 往temp main内写值，匹配度为0
                    tempbook_main_act.cell(yy, 5).value = "未匹配到记录" # 匹配结果
                    tempbook_main_act.cell(yy, 6).value = 0 # 匹配度
                    tempbook_main_act.cell(yy, 7).value = 100 # 可信度
                    tempbook_main.save(tempbook_path_main)

                # 当天没有交易情况，但往后3天有交易情况
                else:
                    df_float = pd.concat([df1, df2, df3])
                    df_float.to_excel(tempbook_path_data2, sheet_name="new_data") # 合并dataframe，写入data2

                    # 根据查询到的往后3天情况，
                    temp_data2_clean = []  # 将dataframe转换为list
                    match_money = []
                    match_date = []
                    match_ser = []
                    temp_table_data2 = xlrd.open_workbook(tempbook_path_data2)
                    sheet_temp_table_data2 = temp_table_data2.sheet_by_name("new_data")
                    temp_data2_rows = sheet_temp_table_data2.nrows
                    x = 1
                    for x in range(1, temp_data2_rows):  # 遍历data2内的内容

                        # 配置环境，读表
                        cell_match_date = 50 # 日期匹配度
                        match_date.append(cell_match_date)
                        cell_temp_data2_number = sheet_temp_table_data2.cell_value(x, 4)
                        cell_temp_data2_ser = sheet_temp_table_data2.cell_value(x, 1)
                        match_ser.append(cell_temp_data2_ser)

                        if cell_temp_data2_number < main_compare_money:
                            # 计算金额浮动值
                            cell_temp_data2_number2 = - (main_compare_money - cell_temp_data2_number) / main_compare_money
                            cell_temp_data2_number2 = round(cell_temp_data2_number2, 4)
                            temp_data2_clean.append(cell_temp_data2_number2)

                            # 计算金额匹配度
                            cell_match_money = (1 + cell_temp_data2_number2) * 100
                            cell_match_money = round(cell_match_money, 2)
                            match_money.append(cell_match_money)

                        elif cell_temp_data2_number > main_compare_money:
                            # 计算金额浮动值
                            cell_temp_data2_number2 = (cell_temp_data2_number - main_compare_money) / main_compare_money
                            cell_temp_data2_number2 = round(cell_temp_data2_number2, 4)
                            temp_data2_clean.append(cell_temp_data2_number2)

                            # 计算金额匹配度
                            cell_match_money = (1 - cell_temp_data2_number2) * 100
                            cell_match_money = round(cell_match_money, 2)
                            match_money.append(cell_match_money)

                    for i in range(0, len(temp_data2_clean)):
                        # 写入临时表
                        j = i + 2
                        tempbook_data_act.cell(j, 1).value = match_ser[i]
                        tempbook_data_act.cell(j, 2).value = temp_data2_clean[i]
                        tempbook_data_act.cell(j, 4).value = match_date[i]
                        tempbook_data_act.cell(j, 5).value = match_money[i]
                    tempbook_data.save(tempbook_path_data)

                    # 读表，排序后提取最高项
                    total = []
                    relia = []
                    # 计算、判断、赋值总体匹配度、可信度
                    temp_table_data = xlrd.open_workbook(tempbook_path_data)  # 读取data
                    sheet_temp_table_data = temp_table_data.sheet_by_name("temp_data")
                    for x in range(1, temp_data2_rows):  # 遍历data2内的内容
                        analysis_data_date = sheet_temp_table_data.cell_value(x, 3)  # 日期匹配度
                        analysis_data_money = sheet_temp_table_data.cell_value(x, 4)  # 金额匹配度
                        # if analysis_data_date == 50:
                        if analysis_data_money < 0:
                            analysis_data_relia = 20
                            analysis_data_total = (50 * 0.3) + (analysis_data_relia * 0.4)
                        elif analysis_data_money > 100:
                            analysis_data_relia = 20
                            analysis_data_total = (50 * 0.3) + (analysis_data_relia * 0.4)
                        elif analysis_data_money < 100 and main_compare_money > 80:
                            analysis_data_relia = 70
                            analysis_data_total = (50 * 0.3) + (analysis_data_money * 0.3) + (analysis_data_relia * 0.4)
                        elif analysis_data_money < 80 and analysis_data_money > 50:
                            analysis_data_relia = 50
                            analysis_data_total = (50 * 0.3) + (analysis_data_money * 0.3) + (analysis_data_relia * 0.4)
                        else:
                            analysis_data_relia = 30
                            analysis_data_total = (50 * 0.3) + (analysis_data_money * 0.3) + (analysis_data_relia * 0.4)
                        total.append(analysis_data_total)
                        relia.append(analysis_data_relia)

                    for i in range(0, len(relia)):
                        # 写入临时表
                        j = i + 2
                        tempbook_data_act.cell(j, 3).value = relia[i]
                        tempbook_data_act.cell(j, 6).value = total[i]
                    tempbook_data.save(tempbook_path_data)

                    # 综合排序：可信度、总体匹配度
                    analysis_range_date = pd.read_excel(tempbook_path_data)
                    analysis_range_date.sort_values(by=['可信度', '总体匹配度'], inplace=True, ascending=False)
                    analysis_range_date.to_excel(tempbook_path_data, sheet_name="temp_data")
                    analysis_total = pd.DataFrame(analysis_range_date)

                    # 筛选总体匹配度中最高的值
                    analysis_total_max = analysis_total['总体匹配度'].max()
                    analysis_total_max = round(analysis_total_max, 3)  # 保留3位小数

                    # 设置环境（设置浮动范围20%，读表data）
                    analysis_total_max_range = analysis_total_max * 0.8
                    temp_table_data_pd = pd.read_excel(tempbook_path_data)  # 使用pd读取data
                    temp_table_data_xlrd = xlrd.open_workbook(tempbook_path_data)  # 使用xlrd读取data
                    sheet_temp_table_data = temp_table_data_xlrd.sheet_by_name("temp_data")  # 使用xlrd选表
                    total_max = []

                    # 从data中筛选数据后写入data3。
                    # 逐行读取值。如果总体匹配度等于最高值，就写入列表data3；低于最高值的20%就不做考虑
                    for y in range(1, temp_data2_rows):  # 取匹配值（超过80%匹配度）以后写入新表data3，再追加到df中。
                        analysis_data_total_cell = sheet_temp_table_data.cell_value(y, 6)  # 总体匹配度
                        match_to_data3 = temp_table_data_pd[(temp_table_data_pd['总体匹配度'] > analysis_total_max_range)]
                        match_to_data3 = pd.DataFrame(match_to_data3)
                        match_to_data3.to_excel(tempbook_path_data3, sheet_name="temp_data")
                        match_to_data3 = match_to_data3['总体匹配度']

                    # 读表data3。
                    temp_table_data3_pd = pd.read_excel(tempbook_path_data3)  # 使用pd读取data3
                    temp_table_data3_xlrd = xlrd.open_workbook(tempbook_path_data3)  # 使用xlrd读取data3
                    sheet_temp_table_data3 = temp_table_data3_xlrd.sheet_by_name("temp_data")  # 使用xlrd选表
                    data_match = temp_table_data3_pd[temp_table_data3_pd['总体匹配度'] == analysis_total_max]
                    len_max = len(data_match)
                    len_total = sheet_temp_table_data3.nrows

                    # 判断有几个最高值。只有1个，直接输出。有多个，合并输出。要细分出多个近似值
                    if len_max == 1 and len_total == 1:  # 只有1个值，就输出该值
                        # 读表data3。
                        total_max.append(match_to_data3)
                        temp_table_data3_xlrd = xlrd.open_workbook(tempbook_path_data3)  # 用xlrd读取data3
                        sheet_temp_data3 = temp_table_data3_xlrd.sheet_by_name("temp_data")

                        # 提取数据
                        temp_data3_ser = sheet_temp_data3.cell_value(1, 2)  # 序号
                        temp_data3_relia = int(sheet_temp_data3.cell_value(1, 4))  # 可信度
                        temp_data3_match = sheet_temp_data3.cell_value(1, 7)  # 匹配度

                        # 数据处理
                        status_output = "延后3日匹配"

                        match_output = str(round(np.mean(temp_data3_match), 2)) + "%"
                        relia_output = str(np.mean(temp_data3_relia)) + "%"
                        return_value_ser = "【序号】" + str(int(temp_data3_ser))
                        details_output = "当前值信息：" + return_value_ser

                        # 写表
                        tempbook_main_act.cell(yy, 5).value = status_output  # 输出匹配情况
                        tempbook_main_act.cell(yy, 6).value = match_output  # 输出总匹配度
                        tempbook_main_act.cell(yy, 7).value = relia_output  # 输出总可信度
                        tempbook_main_act.cell(yy, 8).value = details_output  # 输出匹配明细
                        tempbook_main.save(tempbook_path_main)

                    elif len_max == 1 and len_total < 6:  # 有1个最高值，且有多个近似值（不超过5个）
                        # 读表data3。
                        total_max.append(match_to_data3)
                        temp_table_data3_xlrd = xlrd.open_workbook(tempbook_path_data3)  # 用xlrd读取data3
                        sheet_temp_data3 = temp_table_data3_xlrd.sheet_by_name("temp_data")

                        # 提取数据
                        temp_data3_relia = int(sheet_temp_data3.cell_value(1, 4))  # 可信度
                        temp_data3_match = sheet_temp_data3.cell_value(1, 7)  # 匹配度

                        # 提取数据
                        return_value_all = []
                        relia_all = []
                        match_all = []
                        for z in range(1, len_total):
                            temp_data3_ser = sheet_temp_data3.cell_value(z, 2)  # 序号
                            temp_data3_relia = int(sheet_temp_data3.cell_value(z, 4))  # 可信度
                            temp_data3_match = sheet_temp_data3.cell_value(z, 7)  # 匹配度
                            relia_all.append(temp_data3_relia)
                            match_all.append(temp_data3_match)

                            # 数据处理
                            return_value_ser = "值" + str(z) + "：【序号】" + str(int(temp_data3_ser))
                            return_value_relia = "【可信度】" + str(temp_data3_relia)
                            return_value_match = "【匹配度】" + str(round(temp_data3_match,3)) + "%。"
                            return_value = return_value_ser + return_value_relia + return_value_match
                            return_value_all.append(return_value)
                        status_output = "多项近似匹配"
                        match_output = str(round(np.mean(match_all), 2)) + "%"
                        relia_output = str(np.mean(relia_all)) + "%"
                        str1 = ' '
                        zhi = str1.join(return_value_all)
                        return_value_output = "共匹配到 " + str(len_total - 1) + " 个值。" + str(zhi)

                        # 提交数据
                        tempbook_main_act.cell(yy, 5).value = status_output  # 输出匹配情况
                        tempbook_main_act.cell(yy, 6).value = match_output  # 输出总匹配度
                        tempbook_main_act.cell(yy, 7).value = relia_output  # 输出总可信度
                        tempbook_main_act.cell(yy, 8).value = return_value_output  # 输出匹配明细
                        tempbook_main.save(tempbook_path_main)

                    elif len_max == 1 and len_total > 5:
                        # 读表data3。
                        total_max.append(match_to_data3)
                        temp_table_data3_xlrd = xlrd.open_workbook(tempbook_path_data3)  # 用xlrd读取data3
                        sheet_temp_data3 = temp_table_data3_xlrd.sheet_by_name("temp_data")

                        # 提取数据
                        temp_data3_relia = int(sheet_temp_data3.cell_value(1, 4))  # 可信度
                        temp_data3_match = sheet_temp_data3.cell_value(1, 7)  # 匹配度

                        # 提取数据
                        return_value_all = []
                        relia_all = []
                        match_all = []
                        for z in range(1, 6):
                            temp_data3_ser = sheet_temp_data3.cell_value(z, 2)  # 序号
                            temp_data3_relia = int(sheet_temp_data3.cell_value(z, 4))  # 可信度
                            temp_data3_match = sheet_temp_data3.cell_value(z, 7)  # 匹配度
                            relia_all.append(temp_data3_relia)
                            match_all.append(temp_data3_match)

                            # 数据处理
                            return_value_ser = "值" + str(z) + "：【序号】" + str(int(temp_data3_ser))
                            return_value_relia = "【可信度】" + str(temp_data3_relia)
                            return_value_match = "【匹配度】" + str(round(temp_data3_match, 3)) + "%。"
                            return_value = return_value_ser + return_value_relia + return_value_match
                            return_value_all.append(return_value)
                        status_output = "多项近似匹配"
                        match_output = str(round(np.mean(match_all), 2)) + "%"
                        relia_output = str(np.mean(relia_all)) + "%"
                        str1 = ' '
                        zhi = str1.join(return_value_all)
                        return_value_output = "共匹配到 " + str(len_total - 1) + " 个值。" + str(zhi)

                        # 提交数据
                        tempbook_main_act.cell(yy, 5).value = status_output  # 输出匹配情况
                        tempbook_main_act.cell(yy, 6).value = match_output  # 输出总匹配度
                        tempbook_main_act.cell(yy, 7).value = relia_output  # 输出总可信度
                        tempbook_main_act.cell(yy, 8).value = return_value_output  # 输出匹配明细
                        tempbook_main.save(tempbook_path_main)

                    elif len_max > 1:  # 有多个最高值，合并输出。

                        # 将data3数值拼接、计算后写入temp main表
                        # 读表
                        total_max.append(match_to_data3)
                        temp_table_data3_xlrd = xlrd.open_workbook(tempbook_path_data3)  # 用xlrd读取data3
                        sheet_temp_data3 = temp_table_data3_xlrd.sheet_by_name("temp_data")

                        # 提取数据
                        return_value_all = []
                        relia_all = []
                        match_all = []
                        for z in range(1, len_total):
                            temp_data3_ser = sheet_temp_data3.cell_value(z, 2)  # 序号
                            temp_data3_relia = int(sheet_temp_data3.cell_value(z, 4))  # 可信度
                            temp_data3_match = sheet_temp_data3.cell_value(z, 7)  # 匹配度
                            relia_all.append(temp_data3_relia)
                            match_all.append(temp_data3_match)

                            # 数据处理
                            return_value_ser = "值" + str(z) + "：【序号】" + str(int(temp_data3_ser))
                            return_value_relia = "【可信度】" + str(temp_data3_relia)
                            return_value_match = "【匹配度】" + str(round(temp_data3_match,3)) + "%。"
                            return_value = return_value_ser + return_value_relia + return_value_match
                            return_value_all.append(return_value)
                        status_output = "多项近似匹配"
                        match_output = str(round(np.mean(match_all), 2)) + "%"
                        relia_output = str(np.mean(relia_all)) + "%"
                        str1 = ' '
                        zhi = str1.join(return_value_all)
                        return_value_output = "共匹配到 " + str(len_total - 1) + " 个值。" + str(zhi)

                        # 提交数据
                        tempbook_main_act.cell(yy, 5).value = status_output  # 输出匹配情况
                        tempbook_main_act.cell(yy, 6).value = match_output  # 输出总匹配度
                        tempbook_main_act.cell(yy, 7).value = relia_output  # 输出总可信度
                        tempbook_main_act.cell(yy, 8).value = return_value_output  # 输出匹配明细
                        tempbook_main.save(tempbook_path_main)

            # 能查到日期
            else:
                df.to_excel(tempbook_path_data2, sheet_name="new_data")  # 写入data2
                temp_data2_clean = []
                match_money = []
                match_date = []
                match_ser = []
                temp_table_data2 = xlrd.open_workbook(tempbook_path_data2) # 读取data2
                sheet_temp_table_data2 = temp_table_data2.sheet_by_name("new_data")
                temp_data2_rows = sheet_temp_table_data2.nrows
                x = 1
                for x in range(1, temp_data2_rows):  # 遍历data2内的内容
                    cell_temp_data2_number = sheet_temp_table_data2.cell_value(x, 4)
                    cell_temp_data2_ser = sheet_temp_table_data2.cell_value(x, 1)

                    cell_match_date = 100 # 日期匹配度
                    match_date.append(cell_match_date)
                    match_ser.append(cell_temp_data2_ser)

                    if cell_temp_data2_number < main_compare_money:
                        # 计算金额浮动值
                        cell_temp_data2_number2 = - (main_compare_money - cell_temp_data2_number) / main_compare_money
                        cell_temp_data2_number2 = round(cell_temp_data2_number2, 4)
                        temp_data2_clean.append(cell_temp_data2_number2)

                        # 计算金额匹配度
                        cell_match_money = (1 + cell_temp_data2_number2) * 100
                        cell_match_money = round(cell_match_money, 2)
                        match_money.append(cell_match_money)

                    elif cell_temp_data2_number > main_compare_money:
                        # 计算金额浮动值
                        cell_temp_data2_number2 = (cell_temp_data2_number - main_compare_money) / main_compare_money
                        cell_temp_data2_number2 = round(cell_temp_data2_number2, 4)
                        temp_data2_clean.append(cell_temp_data2_number2)

                        # 计算金额匹配度
                        cell_match_money = (1 - cell_temp_data2_number2) * 100
                        cell_match_money = round(cell_match_money, 2)
                        match_money.append(cell_match_money)

                for i in range(0, len(temp_data2_clean)):
                    # 写入临时表
                    j = i + 2
                    tempbook_data_act.cell(j, 1).value = match_ser[i]
                    tempbook_data_act.cell(j, 2).value = temp_data2_clean[i]
                    tempbook_data_act.cell(j, 4).value = match_date[i]
                    tempbook_data_act.cell(j, 5).value = match_money[i]
                tempbook_data.save(tempbook_path_data)

                total = []
                relia = []
                # 计算、判断、赋值总体匹配度、可信度
                temp_table_data = xlrd.open_workbook(tempbook_path_data)  # 读取data
                sheet_temp_table_data = temp_table_data.sheet_by_name("temp_data")
                for x in range(1, temp_data2_rows):  # 遍历data2内的内容
                    analysis_data_date = sheet_temp_table_data.cell_value(x, 3) # 日期匹配度
                    analysis_data_money = sheet_temp_table_data.cell_value(x, 4) # 金额匹配度
                    if analysis_data_date == 50:
                        if analysis_data_money < 0:
                            analysis_data_relia = 20
                            analysis_data_total = (50 * 0.3) + (analysis_data_relia * 0.4)
                            total.append(analysis_data_total)
                            relia.append(analysis_data_relia)
                        elif main_compare_money > 100:
                            analysis_data_relia = 20
                            analysis_data_total = (50 * 0.3) + (analysis_data_relia * 0.4)
                            total.append(analysis_data_total)
                            relia.append(analysis_data_relia)
                        else:
                            analysis_data_relia = 30
                            analysis_data_total = (50 * 0.3) + (analysis_data_money * 0.3) + (analysis_data_relia * 0.4)
                            total.append(analysis_data_total)
                            relia.append(analysis_data_relia)

                    elif analysis_data_date == 100:
                        if analysis_data_money < 0:
                            analysis_data_relia = 80
                            analysis_data_total = (50 * 0.3) + (analysis_data_relia * 0.4)
                            total.append(analysis_data_total)
                            relia.append(analysis_data_relia)
                        elif analysis_data_money > 100:
                            analysis_data_relia = 80
                            analysis_data_total = (50 * 0.3) + (analysis_data_relia * 0.4)
                            total.append(analysis_data_total)
                            relia.append(analysis_data_relia)
                        else:
                            analysis_data_relia = 90
                            analysis_data_total = (50 * 0.3) + (analysis_data_money * 0.3) + (analysis_data_relia * 0.4)
                            total.append(analysis_data_total)
                            relia.append(analysis_data_relia)

                for i in range(0, len(relia)):
                    # 写入临时表
                    j = i + 2
                    tempbook_data_act.cell(j, 3).value = relia[i]
                    tempbook_data_act.cell(j, 6).value = total[i]
                tempbook_data.save(tempbook_path_data)

                # 综合排序：可信度、总体匹配度
                analysis_range_date = pd.read_excel(tempbook_path_data)
                analysis_range_date.sort_values(by=['可信度','总体匹配度'],inplace=True,ascending=False)
                analysis_range_date.to_excel(tempbook_path_data, sheet_name="temp_data")
                analysis_total = pd.DataFrame(analysis_range_date)

                # 筛选总体匹配度中最高的值
                analysis_total_max = analysis_total['总体匹配度'].max()
                analysis_total_max = round(analysis_total_max, 3) # 保留3位小数

                # 设置环境（设置浮动范围20%，读表data）
                analysis_total_max_range = analysis_total_max * 0.8
                temp_table_data_pd = pd.read_excel(tempbook_path_data)  # 使用pd读取data
                temp_table_data_xlrd = xlrd.open_workbook(tempbook_path_data)  # 使用xlrd读取data
                sheet_temp_table_data = temp_table_data_xlrd.sheet_by_name("temp_data") # 使用xlrd选表
                total_max = []

                # 从data中筛选数据后写入data3。
                # 逐行读取值。如果总体匹配度等于最高值，就写入列表data3；低于最高值的20%就不做考虑
                for y in range(1, temp_data2_rows): # 取匹配值（超过80%匹配度）以后写入新表data3，再追加到df中。
                    analysis_data_total_cell = sheet_temp_table_data.cell_value(y, 6) # 总体匹配度
                    match_to_data3 = temp_table_data_pd[(temp_table_data_pd['总体匹配度'] > analysis_total_max_range)]
                    match_to_data3 = pd.DataFrame(match_to_data3)
                    match_to_data3.to_excel(tempbook_path_data3, sheet_name="temp_data")
                    match_to_data3 = match_to_data3['总体匹配度']

                # 读表data3。
                temp_table_data3_pd = pd.read_excel(tempbook_path_data3)  # 使用pd读取data3
                temp_table_data3_xlrd = xlrd.open_workbook(tempbook_path_data3)  # 使用xlrd读取data3
                sheet_temp_table_data3 = temp_table_data3_xlrd.sheet_by_name("temp_data") # 使用xlrd选表
                data_match = temp_table_data3_pd[temp_table_data3_pd['总体匹配度'] == analysis_total_max]
                len_max = len(data_match)


                # 判断有几个最高值。只有1个，直接输出。有多个，合并输出
                if len_max == 1: # 只有1个值，就输出该值

                    # 读表data3。
                    total_max.append(match_to_data3)
                    temp_table_data3_xlrd = xlrd.open_workbook(tempbook_path_data3) # 用xlrd读取data3
                    sheet_temp_data3 = temp_table_data3_xlrd.sheet_by_name("temp_data")

                    # 提取数据
                    temp_data3_ser = sheet_temp_data3.cell_value(1, 2)  # 序号
                    temp_data3_relia = int(sheet_temp_data3.cell_value(1, 4))  # 可信度
                    temp_data3_match = sheet_temp_data3.cell_value(1, 7)  # 匹配度

                    # 数据处理
                    # 判断，如果匹配度为100，输出为【完整匹配】
                    if analysis_total_max == 100:
                        status_output = "已匹配"
                    elif analysis_total_max < 100:
                        status_output = "单项近似匹配"

                    match_output = str(round(np.mean(temp_data3_match),2)) + "%"
                    relia_output = str(np.mean(temp_data3_relia)) + "%"
                    return_value_ser = "【序号】" + str(int(temp_data3_ser))
                    details_output = "当前值信息：" + return_value_ser



                    # 写表
                    tempbook_main_act.cell(yy, 5).value = status_output  # 输出匹配情况
                    tempbook_main_act.cell(yy, 6).value = match_output # 输出总匹配度
                    tempbook_main_act.cell(yy, 7).value = relia_output # 输出总可信度
                    tempbook_main_act.cell(yy, 8).value = details_output # 输出匹配明细
                    tempbook_main.save(tempbook_path_main)

                elif len_max > 1: # 有多个值，合并输出

                    # 将data3数值拼接、计算后写入temp main表
                    total_max.append(match_to_data3)
                    temp_table_data3_xlrd = xlrd.open_workbook(tempbook_path_data3) # 用xlrd读取data3
                    sheet_temp_data3 = temp_table_data3_xlrd.sheet_by_name("temp_data")
                    temp_data3_rows = sheet_temp_data3.nrows
                    return_value_all = []
                    relia_all = []
                    match_all = []
                    for z in range(1, temp_data3_rows):
                        temp_data3_ser = sheet_temp_data3.cell_value(z,2) # 序号
                        temp_data3_relia = int(sheet_temp_data3.cell_value(z,4)) # 可信度
                        temp_data3_match = sheet_temp_data3.cell_value(z,7) # 匹配度
                        relia_all.append(temp_data3_relia)
                        match_all.append(temp_data3_match)
                        l = z-1
                        return_value_ser = "值" + str(z) + "：【序号】" + str(int(temp_data3_ser))
                        return_value_relia = "【可信度】" + str(temp_data3_relia)
                        return_value_match = "【匹配度】" + str(round(temp_data3_match, 3)) + "%。"
                        return_value = return_value_ser + return_value_relia + return_value_match
                        return_value_all.append(return_value)
                    status_output = "多项近似匹配"
                    match_output = str(round(np.mean(match_all),2)) + "%"
                    relia_output = str(np.mean(relia_all)) + "%"
                    tempbook_main_act.cell(yy, 5).value = status_output # 输出匹配情况
                    tempbook_main_act.cell(yy, 6).value = match_output # 输出总匹配度
                    tempbook_main_act.cell(yy, 7).value = relia_output # 输出总可信度

                    str1 = ' '
                    zhi = str1.join(return_value_all)
                    return_value_output =  "共匹配到 " + str(temp_data3_rows - 1) + " 个值。" + str(zhi)
                    tempbook_main_act.cell(yy, 8).value = return_value_output # 输出匹配明细
                    tempbook_main.save(tempbook_path_main)


            # 读取temp data，计算匹配情况，写入temp main。
            # temp_table_data = xlrd.open_workbook(tempbook_path_data)  # 读取data
            # sheet_temp_table_data = temp_table_data.sheet_by_name("temp_data")

            # 临时
            # sub_compare = pd.read_excel(tempbook_path_sub)
            # sub_compare_date_find = sub_compare[(sub_compare['交易日期'] == main_compare_date) & (sub_compare['交易金额'] > 0)]
            # df = pd.DataFrame(sub_compare_date_find)

            # 匹配总体情况

            # print("已完成",xx,"/",main_rows)
            bar()
        except Exception as e:
            print("遇到错误：", e)




